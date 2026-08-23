import sys
import os
import csv
from datetime import datetime, timedelta
from collections import defaultdict
import pandas as pd
from PyQt6.QtWidgets import (
    QApplication, QWidget, QLabel, QLineEdit, QPushButton, 
    QVBoxLayout, QHBoxLayout, QGridLayout, QGroupBox,
    QTableWidget, QTableWidgetItem, QDialog, QHeaderView, QMessageBox,
    QTabWidget, QFrame
)
from PyQt6.QtCore import Qt, QTimer
from PyQt6.QtGui import QColor, QBrush, QIntValidator

class KoreanLineEdit(QLineEdit):
    def __init__(self, placeholder="", parent=None):
        super().__init__(parent)
        self.setPlaceholderText(placeholder)
        self._composing = False

    def inputMethodEvent(self, event):
        if event.preeditString():
            self._composing = True
        if event.commitString():
            self._composing = False
        super().inputMethodEvent(event)
        if not event.preeditString():
            self._composing = False

    def isComposing(self):
        return self._composing

    def focusOutEvent(self, event):
        self._composing = False
        super().focusOutEvent(event)


class BookManagerApp(QWidget):
    def __init__(self):
        super().__init__()
        self.df = pd.DataFrame()
        self.loan_status = {}
        self.current_book_data = None
        self.log_file = 'logs.csv'
        self.expected_header = ['날짜', '동작', '학생 이름', '바코드', '제목', '저자', 'AR Level']
        self.MAX_LOAN = 4
        self.OVERDUE_MINUTES = 1   # ★ 테스트용: 1분 연체

        self.init_data()
        self.init_ui()
        self.update_status_bar()

    def init_data(self):
        print("[정보] 도서 데이터베이스를 불러오는 중...")
        try:
            self.df = pd.read_csv('ar_books.csv', sep='\t', encoding='utf-8-sig', dtype=str).fillna('')
            if not self.df.empty and 'Barcode' in self.df.columns:
                self.df = self.df.drop_duplicates(subset=['Barcode'], keep='first')
                self.df = self.df.set_index('Barcode', drop=False)
            print(f"[성공] 총 {len(self.df)}권의 데이터 로드 완료.")
        except Exception as e:
            print(f"[오류] 데이터 로드 실패: {e}")

        if not os.path.isfile(self.log_file):
            with open(self.log_file, 'w', newline='', encoding='utf-8-sig') as f:
                csv.writer(f, quoting=csv.QUOTE_ALL).writerow(self.expected_header)
        else:
            try:
                with open(self.log_file, 'r', encoding='utf-8-sig') as f:
                    reader = csv.reader(f)
                    next(reader, None)
                    for row in reader:
                        if len(row) >= 4:
                            try:
                                loan_date = datetime.strptime(row[0].strip(), '%Y-%m-%d %H:%M:%S')
                            except:
                                loan_date = datetime.now()
                            action = row[1].strip()
                            student_name = row[2].strip()
                            barcode = row[3].strip()
                            if action == "대출":
                                self.loan_status[barcode] = {"name": student_name, "date": loan_date}
                            elif action == "반납" and barcode in self.loan_status:
                                del self.loan_status[barcode]
                print(f"[성공] 기존 대출 기록 복원 완료 (현재 대출 중인 도서: {len(self.loan_status)}권)")
            except Exception as e:
                print(f"[알림] 기존 로그 복원 중 오류 발생: {e}")

    def get_book_info(self, barcode):
        if self.df.empty or barcode not in self.df.index:
            return None
        row = self.df.loc[barcode]
        if isinstance(row, pd.DataFrame):
            row = row.iloc[0]
        return row.to_dict()

    def force_commit_ime(self):
        focused = QApplication.focusWidget()
        if focused and isinstance(focused, KoreanLineEdit):
            focused.clearFocus()
            QApplication.processEvents()
            focused.setFocus()
            QApplication.processEvents()

    def get_student_id(self):
        name = self.entry_student.text().strip()
        phone = self.entry_phone.text().strip()
        if not name or not phone:
            return None
        return f"{name} ({phone})"

    def get_student_loan_count(self, student_id):
        return sum(1 for info in self.loan_status.values() if info["name"] == student_id)

    def has_borrowed_before(self, student_id, barcode):
        if not os.path.isfile(self.log_file):
            return False
        try:
            with open(self.log_file, 'r', encoding='utf-8-sig') as f:
                reader = csv.reader(f)
                next(reader, None)
                for row in reader:
                    if len(row) >= 4:
                        if row[1].strip() == "대출" and row[2].strip() == student_id and row[3].strip() == barcode:
                            return True
        except:
            pass
        return False

    def get_overdue_count(self):
        count = 0
        for info in self.loan_status.values():
            minutes = (datetime.now() - info["date"]).total_seconds() / 60
            if minutes > self.OVERDUE_MINUTES:
                count += 1
        return count

    def update_status_bar(self):
        total = len(self.loan_status)
        overdue = self.get_overdue_count()
        self.lbl_bottom_status.setText(
            f"📚 현재 대출 중: {total}권    |    ⚠️ 연체: {overdue}권    |    최대 대출: {self.MAX_LOAN}권/인"
        )
        if overdue > 0:
            self.lbl_bottom_status.setStyleSheet(
                "background-color: #742a2a; color: #fed7d7; font-weight: bold; font-size: 13px; padding: 10px; border-radius: 6px;"
            )
        else:
            self.lbl_bottom_status.setStyleSheet(
                "background-color: #2d3748; color: #a0aec0; font-weight: bold; font-size: 13px; padding: 10px; border-radius: 6px;"
            )

    def init_ui(self):
        self.setWindowTitle("📚 도서 관리 시스템 (데모 - 1분 연체)")
        self.resize(580, 720)
        self.setStyleSheet("""
            QWidget { background-color: #1a202c; color: #e2e8f0; font-family: 'Apple SD Gothic Neo', '맑은 고딕', sans-serif; }
            QLineEdit { background-color: #2d3748; color: white; border: 1px solid #4a5568; border-radius: 6px; padding: 8px; font-size: 14px; }
            QPushButton { background-color: #3182ce; color: white; border-radius: 6px; font-weight: bold; font-size: 13px; padding: 10px; }
            QPushButton:hover { background-color: #2b6cb0; }
            QLabel { font-size: 13px; }
            QTabWidget::pane { border: 1px solid #4a5568; border-radius: 8px; background: #1a202c; }
            QTabBar::tab { background: #2d3748; color: #a0aec0; padding: 10px 18px; margin-right: 4px; border-top-left-radius: 8px; border-top-right-radius: 8px; font-weight: bold; font-size: 13px; }
            QTabBar::tab:selected { background: #3182ce; color: white; }
            QGroupBox { border: 1px solid #4a5568; border-radius: 8px; margin-top: 12px; font-weight: bold; color: #90cdf4; padding-top: 10px; }
            QGroupBox::title { subcontrol-origin: margin; left: 12px; padding: 0 6px; }
        """)

        main_layout = QVBoxLayout()
        main_layout.setSpacing(12)
        main_layout.setContentsMargins(14, 14, 14, 14)

        tabs = QTabWidget()
        tabs.setDocumentMode(True)

        # 탭 1: 대출/반납
        tab_loan = QWidget()
        layout_loan = QVBoxLayout(tab_loan)
        layout_loan.setSpacing(12)

        card_bc = QGroupBox("📷 바코드 스캔")
        vbox_bc = QVBoxLayout()
        hbox_bc = QHBoxLayout()
        self.entry_barcode = KoreanLineEdit("바코드 입력 후 Enter")
        self.entry_barcode.returnPressed.connect(self.on_barcode_enter)
        hbox_bc.addWidget(self.entry_barcode)
        btn_bc_ok = QPushButton("확인")
        btn_bc_ok.clicked.connect(self.safe_search)
        hbox_bc.addWidget(btn_bc_ok)
        btn_bc_clear = QPushButton("지우기")
        btn_bc_clear.setStyleSheet("background-color: #4a5568;")
        btn_bc_clear.clicked.connect(lambda: self.entry_barcode.clear())
        hbox_bc.addWidget(btn_bc_clear)
        vbox_bc.addLayout(hbox_bc)
        card_bc.setLayout(vbox_bc)
        layout_loan.addWidget(card_bc)

        card_st = QGroupBox("👤 대여자 정보")
        vbox_st = QVBoxLayout()
        hbox_name = QHBoxLayout()
        hbox_name.addWidget(QLabel("이름:"))
        self.entry_student = KoreanLineEdit("학생 이름")
        self.entry_student.returnPressed.connect(self.on_student_enter)
        hbox_name.addWidget(self.entry_student)
        vbox_st.addLayout(hbox_name)

        hbox_phone = QHBoxLayout()
        hbox_phone.addWidget(QLabel("핸드폰 끝 4자리:"))
        self.entry_phone = KoreanLineEdit("1234")
        self.entry_phone.setMaxLength(4)
        self.entry_phone.setValidator(QIntValidator(0, 9999))
        self.entry_phone.returnPressed.connect(self.on_student_enter)
        hbox_phone.addWidget(self.entry_phone)
        vbox_st.addLayout(hbox_phone)

        hbox_st_btn = QHBoxLayout()
        btn_st_ok = QPushButton("확인")
        btn_st_ok.clicked.connect(self.safe_check)
        hbox_st_btn.addWidget(btn_st_ok)
        btn_st_clear = QPushButton("지우기")
        btn_st_clear.setStyleSheet("background-color: #4a5568;")
        btn_st_clear.clicked.connect(self.clear_student_fields)
        hbox_st_btn.addWidget(btn_st_clear)
        vbox_st.addLayout(hbox_st_btn)
        card_st.setLayout(vbox_st)
        layout_loan.addWidget(card_st)

        card_info = QGroupBox("📖 도서 정보")
        grid = QGridLayout()
        grid.setColumnStretch(1, 1)

        def add_row(r, text):
            k = QLabel(text)
            k.setStyleSheet("color: #a0aec0; font-weight: bold;")
            v = QLabel("-")
            v.setStyleSheet("font-weight: bold; font-size: 13px;")
            grid.addWidget(k, r, 0)
            grid.addWidget(v, r, 1)
            return v

        self.lbl_title = add_row(0, "제      목 :")
        self.lbl_title.setStyleSheet("color: #63b3ed; font-weight: bold;")
        self.lbl_author = add_row(1, "저      자 :")
        self.lbl_ar = add_row(2, "AR Level :")
        self.lbl_lexile = add_row(3, "Lexile    :")
        self.lbl_quiz = add_row(4, "Quiz No  :")
        self.lbl_barcode = add_row(5, "바 코 드  :")
        self.lbl_loan = add_row(6, "대출 상태 :")
        self.lbl_loan.setStyleSheet("color: #48bb78; font-weight: bold;")
        card_info.setLayout(grid)
        layout_loan.addWidget(card_info)

        hbox_btn = QHBoxLayout()
        btn_loan = QPushButton("📥 대출하기")
        btn_loan.setStyleSheet("background-color: #2b6cb0; font-size: 15px; padding: 12px;")
        btn_loan.clicked.connect(lambda: self.safe_process("대출"))
        hbox_btn.addWidget(btn_loan)

        btn_return = QPushButton("📤 반납하기")
        btn_return.setStyleSheet("background-color: #2f855a; font-size: 15px; padding: 12px;")
        btn_return.clicked.connect(lambda: self.safe_process("반납"))
        hbox_btn.addWidget(btn_return)
        layout_loan.addLayout(hbox_btn)

        self.lbl_status = QLabel("준비됨")
        self.lbl_status.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.lbl_status.setStyleSheet("font-weight: bold; color: #fc8181; font-size: 13px; padding: 6px;")
        layout_loan.addWidget(self.lbl_status)

        layout_loan.addStretch()
        tabs.addTab(tab_loan, "📋 대출/반납")

        # 탭 2: 현황 조회
        tab_status = QWidget()
        layout_status = QVBoxLayout(tab_status)
        layout_status.setSpacing(14)

        btn_status = QPushButton("📋 전체 대출 현황 보기")
        btn_status.setStyleSheet("background-color: #805ad5; font-size: 15px; padding: 14px;")
        btn_status.clicked.connect(self.show_loan_status)
        layout_status.addWidget(btn_status)

        btn_overdue = QPushButton("⚠️ 연체자 목록 보기")
        btn_overdue.setStyleSheet("background-color: #e53e3e; font-size: 15px; padding: 14px;")
        btn_overdue.clicked.connect(self.show_overdue_list)
        layout_status.addWidget(btn_overdue)

        btn_ar_available = QPushButton("📊 AR Level별 대여 가능 현황")
        btn_ar_available.setStyleSheet("background-color: #38a169; font-size: 15px; padding: 14px;")
        btn_ar_available.clicked.connect(self.show_available_by_ar)
        layout_status.addWidget(btn_ar_available)

        layout_status.addStretch()
        tabs.addTab(tab_status, "📊 현황 조회")

        # 탭 3: 도서 관리
        tab_manage = QWidget()
        layout_manage = QVBoxLayout(tab_manage)
        layout_manage.setSpacing(14)

        btn_add_book = QPushButton("➕ 새 도서 등록")
        btn_add_book.setStyleSheet("background-color: #dd6b20; font-size: 15px; padding: 14px;")
        btn_add_book.clicked.connect(self.show_add_book_dialog)
        layout_manage.addWidget(btn_add_book)

        btn_delete_book = QPushButton("🗑️ 도서 삭제")
        btn_delete_book.setStyleSheet("background-color: #c53030; font-size: 15px; padding: 14px;")
        btn_delete_book.clicked.connect(self.show_delete_book_dialog)
        layout_manage.addWidget(btn_delete_book)

        layout_manage.addStretch()
        tabs.addTab(tab_manage, "📚 도서 관리")

        # 탭 4: 엑셀 저장
        tab_excel = QWidget()
        layout_excel = QVBoxLayout(tab_excel)
        layout_excel.setSpacing(14)

        btn_export_loan = QPushButton("📊 대출 현황 엑셀 저장")
        btn_export_loan.setStyleSheet("background-color: #2b6cb0; font-size: 15px; padding: 14px;")
        btn_export_loan.clicked.connect(self.export_loan_status)
        layout_excel.addWidget(btn_export_loan)

        btn_export_log = QPushButton("📜 전체 로그 엑셀 저장")
        btn_export_log.setStyleSheet("background-color: #4a5568; font-size: 15px; padding: 14px;")
        btn_export_log.clicked.connect(self.export_full_log)
        layout_excel.addWidget(btn_export_log)

        layout_excel.addStretch()
        tabs.addTab(tab_excel, "📥 엑셀 저장")

        main_layout.addWidget(tabs)

        self.lbl_bottom_status = QLabel()
        self.lbl_bottom_status.setAlignment(Qt.AlignmentFlag.AlignCenter)
        main_layout.addWidget(self.lbl_bottom_status)

        self.setLayout(main_layout)
        self.entry_barcode.setFocus()

    def clear_student_fields(self):
        self.entry_student.clear()
        self.entry_phone.clear()
        self.entry_student.setFocus()

    def safe_search(self):
        self.force_commit_ime()
        QTimer.singleShot(80, self.search_book)

    def safe_check(self):
        self.force_commit_ime()
        QTimer.singleShot(80, self.check_student)

    def safe_process(self, action):
        self.force_commit_ime()
        QTimer.singleShot(80, lambda: self.process_action(action))

    def on_barcode_enter(self):
        if not self.entry_barcode.isComposing():
            self.safe_search()

    def on_student_enter(self):
        if not self.entry_student.isComposing() and not self.entry_phone.isComposing():
            self.safe_check()

    def search_book(self):
        code = self.entry_barcode.text().strip()
        if not code:
            return
        book = self.get_book_info(code)
        if book:
            self.current_book_data = book
            barcode = book.get('Barcode', '')
            self.lbl_title.setText(book.get('Title', '-'))
            self.lbl_author.setText(book.get('Author', '-'))
            self.lbl_ar.setText(str(book.get('AR Level', '-')))
            self.lbl_lexile.setText(str(book.get('Lexile', '-')))
            self.lbl_quiz.setText(str(book.get('AR Quiz No', '-')))
            self.lbl_barcode.setText(barcode)

            if barcode in self.loan_status:
                info = self.loan_status[barcode]
                self.lbl_loan.setText(f"대출 중 ({info['name']})")
                self.lbl_loan.setStyleSheet("color: #fc8181; font-weight: bold;")
                self.lbl_status.setText(f"[주의] 현재 '{info['name']}'님이 대출 중입니다.")
                self.lbl_status.setStyleSheet("color: #fc8181; font-weight: bold;")
            else:
                self.lbl_loan.setText("대출 가능")
                self.lbl_loan.setStyleSheet("color: #48bb78; font-weight: bold;")
                self.lbl_status.setText("도서 조회 성공. 이름과 핸드폰 끝 4자리를 입력하세요.")
                self.lbl_status.setStyleSheet("color: #63b3ed; font-weight: bold;")
        else:
            self.current_book_data = None
            self.lbl_title.setText("등록되지 않은 도서입니다.")
            self.lbl_author.setText("-")
            self.lbl_ar.setText("-")
            self.lbl_lexile.setText("-")
            self.lbl_quiz.setText("-")
            self.lbl_barcode.setText(code)
            self.lbl_loan.setText("-")
            self.lbl_status.setText("[오류] 데이터베이스에 없는 바코드입니다.")
            self.lbl_status.setStyleSheet("color: #fc8181; font-weight: bold;")
        QTimer.singleShot(50, lambda: self.entry_student.setFocus())

    def check_student(self):
        student_id = self.get_student_id()
        if not student_id:
            self.lbl_status.setText("[경고] 이름과 핸드폰 끝 4자리를 모두 입력하세요!")
            self.lbl_status.setStyleSheet("color: #fc8181; font-weight: bold;")
            return

        count = self.get_student_loan_count(student_id)
        overdue_list = []
        titles = []

        for b, info in self.loan_status.items():
            if info["name"] == student_id:
                book_info = self.get_book_info(b)
                title = book_info.get('Title', '제목없음') if book_info else f"바코드:{b}"
                titles.append(title)
                minutes = (datetime.now() - info["date"]).total_seconds() / 60
                if minutes > self.OVERDUE_MINUTES:
                    overdue_list.append(f"{title} ({int(minutes)}분 연체)")

        msg = f"'{student_id}'님 현재 대출: {count}/{self.MAX_LOAN}권"
        if titles:
            msg += f"\n도서: {', '.join(titles)}"
        if overdue_list:
            msg += f"\n⚠️ 연체 도서: {', '.join(overdue_list)}"
            self.lbl_status.setStyleSheet("color: #fc8181; font-weight: bold;")
        else:
            self.lbl_status.setStyleSheet("color: #48bb78; font-weight: bold;")

        self.lbl_status.setText(msg)

    def process_action(self, action):
        if not self.current_book_data:
            self.lbl_status.setText("[경고] 먼저 도서를 조회해주세요!")
            self.lbl_status.setStyleSheet("color: #fc8181; font-weight: bold;")
            return

        student_id = self.get_student_id()
        if not student_id:
            self.lbl_status.setText("[경고] 이름과 핸드폰 끝 4자리를 모두 입력하세요!")
            self.lbl_status.setStyleSheet("color: #fc8181; font-weight: bold;")
            return

        barcode = self.current_book_data.get('Barcode', '')

        if action == "대출":
            current_count = self.get_student_loan_count(student_id)
            if current_count >= self.MAX_LOAN:
                QMessageBox.warning(self, "대출 제한", f"'{student_id}'님은 이미 {current_count}권을 대출 중입니다.\n최대 {self.MAX_LOAN}권까지 가능합니다.")
                return

            if barcode in self.loan_status:
                self.lbl_status.setText(f"[오류] 이미 '{self.loan_status[barcode]['name']}'님이 대출 중!")
                self.lbl_status.setStyleSheet("color: #fc8181; font-weight: bold;")
                return

            if self.has_borrowed_before(student_id, barcode):
                reply = QMessageBox.information(
                    self, "이전 대여 안내",
                    f"'{student_id}'님은 이전에 이 책을 대여한 적이 있습니다.\n그래도 대출하시겠습니까?",
                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
                )
                if reply == QMessageBox.StandardButton.No:
                    return

        else:
            if barcode not in self.loan_status:
                self.lbl_status.setText("[오류] 대출 기록이 없는 도서입니다!")
                self.lbl_status.setStyleSheet("color: #fc8181; font-weight: bold;")
                return
            if self.loan_status[barcode]["name"] != student_id:
                self.lbl_status.setText(f"[오류] '{self.loan_status[barcode]['name']}'님이 대출한 책입니다!")
                self.lbl_status.setStyleSheet("color: #fc8181; font-weight: bold;")
                return

        try:
            with open(self.log_file, 'a', newline='', encoding='utf-8-sig') as f:
                writer = csv.writer(f, quoting=csv.QUOTE_ALL)
                writer.writerow([
                    datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    action, student_id, barcode,
                    self.current_book_data.get('Title', ''),
                    self.current_book_data.get('Author', ''),
                    str(self.current_book_data.get('AR Level', ''))
                ])
        except Exception as e:
            self.lbl_status.setText(f"[저장 오류] {e}")
            return

        if action == "대출":
            self.loan_status[barcode] = {"name": student_id, "date": datetime.now()}
            self.lbl_loan.setText(f"대출 중 ({student_id})")
            self.lbl_loan.setStyleSheet("color: #fc8181; font-weight: bold;")
        else:
            del self.loan_status[barcode]
            self.lbl_loan.setText("대출 가능 (반납됨)")
            self.lbl_loan.setStyleSheet("color: #48bb78; font-weight: bold;")

        self.lbl_status.setText(f"'{self.current_book_data.get('Title', '')}' → [{action}] 완료! ({student_id})")
        self.lbl_status.setStyleSheet("color: #63b3ed; font-weight: bold;")

        self.entry_barcode.clear()
        self.clear_student_fields()
        self.entry_barcode.setFocus()
        self.update_status_bar()

    def _create_return_button(self, dialog, table, barcode, student_name, update_title_func=None):
        btn = QPushButton("반납하기")
        btn.setObjectName("returnBtn")
        btn.setCursor(Qt.CursorShape.PointingHandCursor)

        def do_return():
            reply = QMessageBox.question(
                dialog, "반납 확인",
                f"이 책을 반납 처리하시겠습니까?\n\n대출자: {student_name}\n바코드: {barcode}",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )
            if reply != QMessageBox.StandardButton.Yes:
                return

            if barcode not in self.loan_status:
                QMessageBox.warning(dialog, "오류", "이미 반납된 도서입니다.")
                return

            try:
                book_info = self.get_book_info(barcode)
                book_title = book_info.get('Title', '') if book_info else ''
                book_author = book_info.get('Author', '') if book_info else ''
                book_ar = str(book_info.get('AR Level', '')) if book_info else ''

                with open(self.log_file, 'a', newline='', encoding='utf-8-sig') as f:
                    writer = csv.writer(f, quoting=csv.QUOTE_ALL)
                    writer.writerow([
                        datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                        "반납", student_name, barcode,
                        book_title, book_author, book_ar
                    ])

                del self.loan_status[barcode]

                for r in range(table.rowCount()):
                    item = table.item(r, 0)
                    if item and item.text() == str(barcode):
                        table.removeRow(r)
                        break

                if update_title_func:
                    update_title_func()

                self.update_status_bar()
                QMessageBox.information(dialog, "반납 완료", f"반납 처리되었습니다.\n\n{book_title}")

            except Exception as e:
                QMessageBox.critical(dialog, "오류", f"반납 처리 중 오류:\n{e}")

        btn.clicked.connect(do_return)
        return btn

    def show_loan_status(self):
        dialog = QDialog(self)
        dialog.setWindowTitle(f"현재 대출 현황 (총 {len(self.loan_status)}권)")
        dialog.resize(1050, 520)
        dialog.setStyleSheet("""
            QDialog { background-color: #1a202c; color: #e2e8f0; }
            QTableWidget { background-color: #2d3748; color: white; gridline-color: #4a5568; font-size: 13px; }
            QHeaderView::section { background-color: #4a5568; color: white; font-weight: bold; padding: 6px; }
            QPushButton { background-color: #3182ce; color: white; border-radius: 4px; font-weight: bold; padding: 6px 10px; }
            QPushButton#returnBtn { background-color: #38a169; }
            QPushButton#returnBtn:hover { background-color: #2f855a; }
        """)

        layout = QVBoxLayout()
        table = QTableWidget()
        table.setColumnCount(8)
        table.setHorizontalHeaderLabels(["바코드", "제목", "저자", "AR Level", "대출자", "대출일", "상태", "반납"])
        table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        table.horizontalHeader().setSectionResizeMode(7, QHeaderView.ResizeMode.Fixed)
        table.setColumnWidth(7, 90)
        table.setSortingEnabled(False)
        table.setAlternatingRowColors(True)
        table.verticalHeader().setVisible(False)

        rows_data = []
        for barcode, info in list(self.loan_status.items()):
            book = self.get_book_info(barcode)
            if book:
                title = book.get('Title', '')
                author = book.get('Author', '')
                ar_level = str(book.get('AR Level', ''))
            else:
                title = "(정보 없음)"
                author = ""
                ar_level = ""

            minutes = (datetime.now() - info["date"]).total_seconds() / 60
            status = f"{int(minutes)}분 경과"
            if minutes > self.OVERDUE_MINUTES:
                status = f"⚠️ 연체 {int(minutes - self.OVERDUE_MINUTES)}분"

            rows_data.append([barcode, title, author, ar_level, info["name"], info["date"].strftime('%Y-%m-%d %H:%M'), status])

        table.setRowCount(len(rows_data))
        for i, row_data in enumerate(rows_data):
            for j, value in enumerate(row_data):
                item = QTableWidgetItem(str(value))
                item.setFlags(item.flags() ^ Qt.ItemFlag.ItemIsEditable)
                if j == 6 and "연체" in str(value):
                    item.setForeground(QBrush(QColor("#fc8181")))
                table.setItem(i, j, item)

            barcode = row_data[0]
            student_name = row_data[4]
            btn = self._create_return_button(
                dialog, table, barcode, student_name,
                update_title_func=lambda: dialog.setWindowTitle(f"현재 대출 현황 (총 {len(self.loan_status)}권)")
            )
            table.setCellWidget(i, 7, btn)

        layout.addWidget(table)
        btn_close = QPushButton("닫기")
        btn_close.clicked.connect(dialog.accept)
        layout.addWidget(btn_close)
        dialog.setLayout(layout)
        dialog.exec()

    def show_overdue_list(self):
        dialog = QDialog(self)
        dialog.setWindowTitle("연체자 목록")
        dialog.resize(1000, 500)
        dialog.setStyleSheet("""
            QDialog { background-color: #1a202c; color: #e2e8f0; }
            QTableWidget { background-color: #2d3748; color: white; gridline-color: #4a5568; font-size: 13px; }
            QHeaderView::section { background-color: #e53e3e; color: white; font-weight: bold; padding: 6px; }
            QPushButton { background-color: #3182ce; color: white; border-radius: 4px; font-weight: bold; padding: 8px; }
            QPushButton#returnBtn { background-color: #38a169; }
            QPushButton#returnBtn:hover { background-color: #2f855a; }
        """)

        layout = QVBoxLayout()

        overdue_items = []
        for barcode, info in self.loan_status.items():
            minutes = (datetime.now() - info["date"]).total_seconds() / 60
            if minutes > self.OVERDUE_MINUTES:
                overdue_items.append((barcode, info, minutes))

        if not overdue_items:
            lbl = QLabel("현재 연체된 도서가 없습니다.")
            lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
            lbl.setStyleSheet("font-size: 16px; color: #68d391; padding: 40px;")
            layout.addWidget(lbl)
            btn_close = QPushButton("닫기")
            btn_close.clicked.connect(dialog.accept)
            layout.addWidget(btn_close)
            dialog.setLayout(layout)
            dialog.exec()
            return

        table = QTableWidget()
        table.setColumnCount(8)
        table.setHorizontalHeaderLabels(["바코드", "제목", "저자", "AR Level", "대출자", "대출일", "연체시간", "반납"])
        table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        table.horizontalHeader().setSectionResizeMode(7, QHeaderView.ResizeMode.Fixed)
        table.setColumnWidth(7, 90)
        table.setSortingEnabled(False)
        table.setAlternatingRowColors(True)
        table.verticalHeader().setVisible(False)

        table.setRowCount(len(overdue_items))
        for i, (barcode, info, minutes) in enumerate(overdue_items):
            book = self.get_book_info(barcode)
            if book:
                title = book.get('Title', '')
                author = book.get('Author', '')
                ar_level = str(book.get('AR Level', ''))
            else:
                title = "(정보 없음)"
                author = ""
                ar_level = ""

            row_data = [
                barcode, title, author, ar_level,
                info["name"], info["date"].strftime('%Y-%m-%d %H:%M'),
                f"{int(minutes - self.OVERDUE_MINUTES)}분 연체"
            ]

            for j, value in enumerate(row_data):
                item = QTableWidgetItem(str(value))
                item.setFlags(item.flags() ^ Qt.ItemFlag.ItemIsEditable)
                if j == 6:
                    item.setForeground(QBrush(QColor("#fc8181")))
                table.setItem(i, j, item)

            btn = self._create_return_button(
                dialog, table, barcode, info["name"],
                update_title_func=lambda: dialog.setWindowTitle(f"연체자 목록 (남은 {table.rowCount()}권)")
            )
            table.setCellWidget(i, 7, btn)

        layout.addWidget(table)
        lbl_count = QLabel(f"총 연체 도서: {len(overdue_items)}권")
        lbl_count.setStyleSheet("font-weight: bold; color: #fc8181; font-size: 14px;")
        layout.addWidget(lbl_count)
        btn_close = QPushButton("닫기")
        btn_close.clicked.connect(dialog.accept)
        layout.addWidget(btn_close)
        dialog.setLayout(layout)
        dialog.exec()

    def show_available_by_ar(self):
        dialog = QDialog(self)
        dialog.setWindowTitle("AR Level별 대여 가능 현황")
        dialog.resize(400, 650)
        dialog.setStyleSheet("""
            QDialog { background-color: #1a202c; color: #e2e8f0; }
            QTableWidget { background-color: #2d3748; color: white; gridline-color: #4a5568; font-size: 14px; }
            QHeaderView::section { background-color: #38a169; color: white; font-weight: bold; padding: 8px; }
            QPushButton { background-color: #3182ce; color: white; border-radius: 4px; font-weight: bold; padding: 8px; }
        """)

        layout = QVBoxLayout()
        table = QTableWidget()
        table.setColumnCount(2)
        table.setHorizontalHeaderLabels(["AR Level", "대여 가능 권수"])
        table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        table.setAlternatingRowColors(True)
        table.verticalHeader().setVisible(False)

        loaned = set(self.loan_status.keys())

        def normalize_ar(value):
            try:
                return f"{float(str(value).strip()):.1f}"
            except:
                return str(value).strip()

        available_count = defaultdict(int)
        for _, row in self.df.iterrows():
            ar = normalize_ar(row.get('AR Level', ''))
            barcode = str(row.get('Barcode', '')).strip()
            if ar and barcode and barcode not in loaned:
                available_count[ar] += 1

        levels = [f"{i/10:.1f}" for i in range(10, 50)]
        table.setRowCount(len(levels))
        for i, level in enumerate(levels):
            count = available_count.get(level, 0)
            item_level = QTableWidgetItem(f"AR {level}")
            item_count = QTableWidgetItem(str(count))
            item_level.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            item_count.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            item_level.setFlags(item_level.flags() ^ Qt.ItemFlag.ItemIsEditable)
            item_count.setFlags(item_count.flags() ^ Qt.ItemFlag.ItemIsEditable)
            if count == 0:
                item_count.setForeground(QBrush(QColor("#a0aec0")))
            else:
                item_count.setForeground(QBrush(QColor("#68d391")))
            table.setItem(i, 0, item_level)
            table.setItem(i, 1, item_count)

        layout.addWidget(table)
        btn_close = QPushButton("닫기")
        btn_close.clicked.connect(dialog.accept)
        layout.addWidget(btn_close)
        dialog.setLayout(layout)
        dialog.exec()

    def show_add_book_dialog(self):
        dialog = QDialog(self)
        dialog.setWindowTitle("새 도서 등록")
        dialog.resize(450, 420)
        dialog.setStyleSheet("""
            QDialog { background-color: #1a202c; color: #e2e8f0; }
            QLineEdit { background-color: #2d3748; color: white; border: 1px solid #4a5568; border-radius: 4px; padding: 8px; font-size: 14px; }
            QLabel { font-size: 13px; font-weight: bold; }
            QPushButton { background-color: #3182ce; color: white; border-radius: 4px; font-weight: bold; padding: 10px; }
        """)

        layout = QVBoxLayout()
        layout.setSpacing(10)

        layout.addWidget(QLabel("바코드"))
        entry_barcode = KoreanLineEdit("바코드 스캔 또는 입력")
        layout.addWidget(entry_barcode)

        layout.addWidget(QLabel("제목"))
        entry_title = KoreanLineEdit("책 제목")
        layout.addWidget(entry_title)

        layout.addWidget(QLabel("저자"))
        entry_author = KoreanLineEdit("저자 이름")
        layout.addWidget(entry_author)

        layout.addWidget(QLabel("AR Level (예: 2.3)"))
        entry_ar = KoreanLineEdit("2.3")
        layout.addWidget(entry_ar)

        layout.addWidget(QLabel("Lexile (선택)"))
        entry_lexile = KoreanLineEdit("450L")
        layout.addWidget(entry_lexile)

        layout.addWidget(QLabel("AR Quiz No (선택)"))
        entry_quiz = KoreanLineEdit("퀴즈 번호")
        layout.addWidget(entry_quiz)

        hbox = QHBoxLayout()
        btn_save = QPushButton("저장")
        btn_cancel = QPushButton("취소")
        btn_cancel.setStyleSheet("background-color: #4a5568;")
        hbox.addWidget(btn_save)
        hbox.addWidget(btn_cancel)
        layout.addLayout(hbox)

        dialog.setLayout(layout)
        entry_barcode.setFocus()

        def save_book():
            barcode = entry_barcode.text().strip()
            title = entry_title.text().strip()
            author = entry_author.text().strip()
            ar_level = entry_ar.text().strip()
            lexile = entry_lexile.text().strip()
            quiz = entry_quiz.text().strip()

            if not barcode or not title or not author or not ar_level:
                QMessageBox.warning(dialog, "입력 오류", "바코드, 제목, 저자, AR Level은 필수입니다.")
                return

            if not self.df.empty and barcode in self.df.index:
                QMessageBox.warning(dialog, "중복 오류", f"이미 등록된 바코드입니다: {barcode}")
                return

            try:
                new_data = {
                    'AR Level': ar_level, 'Lexile': lexile, 'Title': title,
                    'Author': author, 'AR Quiz No': quiz,
                    'IL': '', 'Points': '', 'Fiction': '', 'Rating': '',
                    'Word Count': '', 'RP': '', 'RV': '', 'VP': '', 'LS': '',
                    'Barcode': barcode
                }
                new_row = pd.DataFrame([new_data])
                self.df = pd.concat([self.df, new_row], ignore_index=True)
                self.df = self.df.set_index('Barcode', drop=False)
                self.df.to_csv('ar_books.csv', sep='\t', index=False, encoding='utf-8-sig')
                QMessageBox.information(dialog, "등록 완료", f"새 도서가 등록되었습니다!\n\n제목: {title}\n바코드: {barcode}")
                dialog.accept()
            except Exception as e:
                QMessageBox.critical(dialog, "저장 오류", str(e))

        btn_save.clicked.connect(save_book)
        btn_cancel.clicked.connect(dialog.reject)
        dialog.exec()

    def show_delete_book_dialog(self):
        dialog = QDialog(self)
        dialog.setWindowTitle("도서 삭제")
        dialog.resize(450, 340)
        dialog.setStyleSheet("""
            QDialog { background-color: #1a202c; color: #e2e8f0; }
            QLineEdit { background-color: #2d3748; color: white; border: 1px solid #4a5568; border-radius: 4px; padding: 8px; font-size: 14px; }
            QLabel { font-size: 13px; }
            QPushButton { background-color: #3182ce; color: white; border-radius: 4px; font-weight: bold; padding: 10px; }
        """)

        layout = QVBoxLayout()
        layout.setSpacing(12)

        layout.addWidget(QLabel("삭제할 도서의 바코드를 스캔하거나 입력하세요:"))
        entry_barcode = KoreanLineEdit("바코드 입력")
        layout.addWidget(entry_barcode)

        lbl_info = QLabel("바코드를 입력하면 책 정보가 여기에 표시됩니다.")
        lbl_info.setStyleSheet("color: #a0aec0; padding: 12px; background-color: #2d3748; border-radius: 4px;")
        lbl_info.setWordWrap(True)
        layout.addWidget(lbl_info)

        hbox = QHBoxLayout()
        btn_search = QPushButton("조회")
        btn_delete = QPushButton("삭제")
        btn_delete.setStyleSheet("background-color: #e53e3e;")
        btn_cancel = QPushButton("취소")
        btn_cancel.setStyleSheet("background-color: #4a5568;")
        hbox.addWidget(btn_search)
        hbox.addWidget(btn_delete)
        hbox.addWidget(btn_cancel)
        layout.addLayout(hbox)

        dialog.setLayout(layout)
        entry_barcode.setFocus()

        current_book = {}

        def search_for_delete():
            nonlocal current_book
            code = entry_barcode.text().strip()
            if not code:
                return
            book = self.get_book_info(code)
            if not book:
                lbl_info.setText(f"[오류] 바코드 '{code}'에 해당하는 도서가 없습니다.")
                lbl_info.setStyleSheet("color: #fc8181; padding: 12px; background-color: #2d3748; border-radius: 4px;")
                current_book = {}
                return

            current_book = book
            info_text = (
                f"제목: {book.get('Title', '-')}\n"
                f"저자: {book.get('Author', '-')}\n"
                f"AR Level: {book.get('AR Level', '-')}\n"
                f"바코드: {book.get('Barcode', '-')}"
            )
            if code in self.loan_status:
                info_text += f"\n\n⚠️ 현재 '{self.loan_status[code]['name']}'님이 대출 중입니다!"
                lbl_info.setStyleSheet("color: #fc8181; padding: 12px; background-color: #2d3748; border-radius: 4px;")
            else:
                lbl_info.setStyleSheet("color: #68d391; padding: 12px; background-color: #2d3748; border-radius: 4px;")
            lbl_info.setText(info_text)

        def delete_book():
            if not current_book:
                QMessageBox.warning(dialog, "오류", "먼저 도서를 조회해주세요.")
                return

            barcode = current_book.get('Barcode', '')
            title = current_book.get('Title', '')

            if barcode in self.loan_status:
                QMessageBox.critical(dialog, "삭제 불가",
                    f"이 책은 현재 대출 중입니다.\n반납 후에 삭제할 수 있습니다.\n\n대출자: {self.loan_status[barcode]['name']}")
                return

            reply = QMessageBox.question(
                dialog, "삭제 확인",
                f"정말로 이 도서를 삭제하시겠습니까?\n\n제목: {title}\n바코드: {barcode}\n\n이 작업은 되돌릴 수 없습니다!",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )
            if reply != QMessageBox.StandardButton.Yes:
                return

            try:
                if barcode in self.df.index:
                    self.df = self.df.drop(barcode)
                self.df.to_csv('ar_books.csv', sep='\t', index=False, encoding='utf-8-sig')
                QMessageBox.information(dialog, "삭제 완료", f"도서가 삭제되었습니다.\n\n제목: {title}")
                dialog.accept()
            except Exception as e:
                QMessageBox.critical(dialog, "오류", f"삭제 중 오류 발생:\n{e}")

        btn_search.clicked.connect(search_for_delete)
        entry_barcode.returnPressed.connect(search_for_delete)
        btn_delete.clicked.connect(delete_book)
        btn_cancel.clicked.connect(dialog.reject)
        dialog.exec()

    def export_loan_status(self):
        if not self.loan_status:
            QMessageBox.information(self, "알림", "현재 대출 중인 도서가 없습니다.")
            return
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            wb = Workbook()
            ws = wb.active
            ws.title = "대출현황"
            headers = ["바코드", "제목", "저자", "AR Level", "대출자", "대출일", "경과/연체"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="4A5568")
                cell.alignment = Alignment(horizontal="center")
            for row_idx, (barcode, info) in enumerate(self.loan_status.items(), 2):
                book = self.get_book_info(barcode)
                title = book.get('Title', '') if book else "(정보없음)"
                author = book.get('Author', '') if book else ""
                ar = str(book.get('AR Level', '')) if book else ""
                minutes = (datetime.now() - info["date"]).total_seconds() / 60
                status = f"{int(minutes)}분 경과"
                if minutes > self.OVERDUE_MINUTES:
                    status = f"연체 {int(minutes - self.OVERDUE_MINUTES)}분"
                ws.cell(row=row_idx, column=1, value=barcode)
                ws.cell(row=row_idx, column=2, value=title)
                ws.cell(row=row_idx, column=3, value=author)
                ws.cell(row=row_idx, column=4, value=ar)
                ws.cell(row=row_idx, column=5, value=info["name"])
                ws.cell(row=row_idx, column=6, value=info["date"].strftime('%Y-%m-%d %H:%M'))
                ws.cell(row=row_idx, column=7, value=status)
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws.column_dimensions[column].width = min(max_length + 2, 40)
            filename = f"대출현황_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
            wb.save(filename)
            QMessageBox.information(self, "저장 완료", f"엑셀 파일이 저장되었습니다.\n\n파일명: {filename}")
        except ImportError:
            QMessageBox.critical(self, "오류", "openpyxl 라이브러리가 필요합니다.\n\npip install openpyxl")
        except Exception as e:
            QMessageBox.critical(self, "오류", f"저장 중 오류 발생:\n{e}")

    def export_full_log(self):
        if not os.path.isfile(self.log_file):
            QMessageBox.information(self, "알림", "로그 파일이 없습니다.")
            return
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            wb = Workbook()
            ws = wb.active
            ws.title = "전체로그"
            with open(self.log_file, 'r', encoding='utf-8-sig') as f:
                reader = csv.reader(f)
                headers = next(reader, None)
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill("solid", fgColor="2B6CB0")
                    cell.alignment = Alignment(horizontal="center")
                for row_idx, row in enumerate(reader, 2):
                    for col, value in enumerate(row, 1):
                        ws.cell(row=row_idx, column=col, value=value)
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws.column_dimensions[column].width = min(max_length + 2, 40)
            filename = f"전체로그_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
            wb.save(filename)
            QMessageBox.information(self, "저장 완료", f"엑셀 파일이 저장되었습니다.\n\n파일명: {filename}")
        except ImportError:
            QMessageBox.critical(self, "오류", "openpyxl 라이브러리가 필요합니다.\n\npip install openpyxl")
        except Exception as e:
            QMessageBox.critical(self, "오류", f"저장 중 오류 발생:\n{e}")


if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = BookManagerApp()
    window.show()
    sys.exit(app.exec())
